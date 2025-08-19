import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from ttkbootstrap import Style
from ttkbootstrap.widgets import DateEntry
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)
import threading
import asyncio
import os
import re
from typing import Optional, List
import sqlite3
from datetime import datetime
import os
from docxtpl import DocxTemplate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


class ConcreteDatabase:
    def __init__(self):
        self.conn = sqlite3.connect('concrete.db')
        self.create_tables()
    
    def create_tables(self):
        scripts = [
            """CREATE TABLE IF NOT EXISTS organizations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                contact TEXT,
                phone TEXT)""",
            """CREATE TABLE IF NOT EXISTS objects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                org_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                address TEXT,
                FOREIGN KEY (org_id) REFERENCES organizations(id))""",
            """CREATE TABLE IF NOT EXISTS constructions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                object_id INTEGER NOT NULL,
                pour_date TEXT NOT NULL,
                element TEXT,
                concrete_class TEXT,
                frost_resistance TEXT,
                water_resistance TEXT,
                supplier TEXT,
                concrete_passport TEXT,
                volume_concrete REAL,
                cubes_count INTEGER,
                cones_count INTEGER,
                slump TEXT,
                temperature TEXT,
                temp_measurements INTEGER,
                executor TEXT,
                act_number TEXT,
                request_number TEXT,
                FOREIGN KEY (object_id) REFERENCES objects(id))"""
        ]
        cursor = self.conn.cursor()
        for script in scripts:
            cursor.execute(script)
        # Миграция: добавляем колонку 'invoice' в constructions, если отсутствует
        cursor.execute("PRAGMA table_info(constructions)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        if 'invoice' not in column_names:
            cursor.execute("ALTER TABLE constructions ADD COLUMN invoice TEXT")
        self.conn.commit()


# =================== Telegram Bot Service ===================
TELEGRAM_BOT_TOKEN = os.getenv(
    'TELEGRAM_BOT_TOKEN',
    '7619596833:AAEtEBVEcQeyevk61-kZdXvpZp1skfdzomA'
)

class TelegramBotService:
    """Background Telegram bot that guides user to add a new construction record."""

    # Conversation states
    (ACTION, ORG, OBJ, CLASS, FROST, WATER, ELEMENT, SUPPLIER, PASSPORT,
     CUBES, CONES, SLUMP, VOLUME, TEMP, TEMP_MEAS, EXECUTOR, ACT, DATE) = range(18)

    def __init__(self, token: str, db_path: str = 'concrete.db'):
        self.token = token
        self.db_path = db_path
        self.thread: Optional[threading.Thread] = None

    def is_running(self) -> bool:
        return self.thread is not None and self.thread.is_alive()

    def start(self) -> None:
        if self.is_running():
            return
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def _run(self) -> None:
        print("[TG] Building application...")
        application = ApplicationBuilder().token(self.token).build()

        # Local DB connection for this thread
        try:
            db_conn = sqlite3.connect(self.db_path)
        except Exception as e:
            print(f"[TG] DB connect error: {e}")
            return

        def fetchall(query: str, params: tuple = ()):
            cur = db_conn.cursor()
            cur.execute(query, params)
            return cur.fetchall()

        def fetch_distinct(column: str) -> List[str]:
            rows = fetchall(
                f"SELECT DISTINCT {column} FROM constructions WHERE {column} IS NOT NULL AND {column} != '' ORDER BY {column}"
            )
            return [r[0] for r in rows]

        async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
            keyboard = [[InlineKeyboardButton("Хочу добавить контроль", callback_data="ACTION:ADD")]]
            await update.effective_message.reply_text(
                "Че надо?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.ACTION

        async def action_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            if query.data == "ACTION:ADD":
                # Ask for organization
                orgs = fetchall("SELECT id, name FROM organizations ORDER BY name")
                if not orgs:
                    await query.edit_message_text("Нет организаций в базе.")
                    return ConversationHandler.END
                keyboard = []
                row = []
                for oid, name in orgs:
                    row.append(InlineKeyboardButton(name, callback_data=f"ORG:{oid}"))
                    if len(row) == 2:
                        keyboard.append(row)
                        row = []
                if row:
                    keyboard.append(row)
                await query.edit_message_text(
                    "Какая организация?",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                return self.ORG
            return ConversationHandler.END

        async def org_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            m = re.match(r"ORG:(\d+)", query.data)
            if not m:
                await query.edit_message_text("Неверный выбор организации")
                return ConversationHandler.END
            org_id = int(m.group(1))
            context.user_data['org_id'] = org_id

            objs = fetchall("SELECT id, name FROM objects WHERE org_id=? ORDER BY name", (org_id,))
            if not objs:
                await query.edit_message_text("У организации нет объектов.")
                return ConversationHandler.END
            keyboard = []
            row = []
            for oid, name in objs:
                row.append(InlineKeyboardButton(name, callback_data=f"OBJ:{oid}"))
                if len(row) == 2:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            await query.edit_message_text(
                "Какой объект?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.OBJ

        async def obj_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            m = re.match(r"OBJ:(\d+)", query.data)
            if not m:
                await query.edit_message_text("Неверный выбор объекта")
                return ConversationHandler.END
            object_id = int(m.group(1))
            context.user_data['object_id'] = object_id

            classes = fetch_distinct('concrete_class')
            if not classes:
                classes = ["B15", "B20", "B25", "B30"]
            keyboard = []
            row = []
            for val in classes:
                row.append(InlineKeyboardButton(val, callback_data=f"CLASS:{val}"))
                if len(row) == 3:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            await query.edit_message_text(
                "Класс бетона?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.CLASS

        async def class_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            context.user_data['concrete_class'] = query.data.split(":", 1)[1]

            frosts = fetch_distinct('frost_resistance')
            if not frosts:
                frosts = ["F100", "F150", "F200"]
            keyboard = []
            row = []
            for val in frosts:
                row.append(InlineKeyboardButton(val, callback_data=f"FROST:{val}"))
                if len(row) == 3:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            await query.edit_message_text(
                "Морозостойкость?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.FROST

        async def frost_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            context.user_data['frost_resistance'] = query.data.split(":", 1)[1]

            waters = fetch_distinct('water_resistance')
            if not waters:
                waters = ["W4", "W6", "W8", "W10"]
            keyboard = []
            row = []
            for val in waters:
                row.append(InlineKeyboardButton(val, callback_data=f"WATER:{val}"))
                if len(row) == 3:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            await query.edit_message_text(
                "Водопроницаемость?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.WATER

        async def water_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            context.user_data['water_resistance'] = query.data.split(":", 1)[1]
            await query.edit_message_text("Конструктив? (введите текст)")
            return self.ELEMENT

        async def element_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            context.user_data['element'] = update.message.text.strip()
            suppliers = fetch_distinct('supplier')
            if not suppliers:
                suppliers = ["Неизвестно"]
            keyboard = []
            row = []
            for val in suppliers:
                row.append(InlineKeyboardButton(val, callback_data=f"SUPPLIER:{val}"))
                if len(row) == 2:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            await update.message.reply_text(
                "Поставщик?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.SUPPLIER

        async def supplier_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            context.user_data['supplier'] = query.data.split(":", 1)[1]
            await query.edit_message_text("Паспорт? (введите текст)")
            return self.PASSPORT

        async def passport_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            context.user_data['concrete_passport'] = update.message.text.strip()
            await update.message.reply_text("Количество кубиков? (число)")
            return self.CUBES

        async def cubes_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            try:
                context.user_data['cubes_count'] = int(update.message.text.strip())
            except Exception:
                await update.message.reply_text("Введите целое число для кубиков")
                return self.CUBES
            await update.message.reply_text("Количество конусов? (число)")
            return self.CONES

        async def cones_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            try:
                context.user_data['cones_count'] = int(update.message.text.strip())
            except Exception:
                await update.message.reply_text("Введите целое число для конусов")
                return self.CONES
            await update.message.reply_text("Просадка? (введите текст, например 10)")
            return self.SLUMP

        async def slump_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            context.user_data['slump'] = update.message.text.strip()
            await update.message.reply_text("Объем бетонной смеси? (число, можно с точкой)")
            return self.VOLUME

        async def volume_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            try:
                context.user_data['volume_concrete'] = float(update.message.text.strip())
            except Exception:
                await update.message.reply_text("Введите число для объема")
                return self.VOLUME
            await update.message.reply_text("Температура? (введите текст, например 12)")
            return self.TEMP

        async def temp_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            context.user_data['temperature'] = update.message.text.strip()
            await update.message.reply_text("Сколько замеров темп.? (целое число)")
            return self.TEMP_MEAS

        async def temp_meas_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            try:
                context.user_data['temp_measurements'] = int(update.message.text.strip())
            except Exception:
                await update.message.reply_text("Введите целое число для замеров")
                return self.TEMP_MEAS

            executors = fetch_distinct('executor')
            if not executors:
                executors = ["Исполнитель"]
            keyboard = []
            row = []
            for val in executors:
                row.append(InlineKeyboardButton(val, callback_data=f"EXECUTOR:{val}"))
                if len(row) == 2:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            await update.message.reply_text(
                "Исполнитель?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.EXECUTOR

        async def executor_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            context.user_data['executor'] = query.data.split(":", 1)[1]
            await query.edit_message_text("Как назвать Акт? (введите текст)")
            return self.ACT

        async def act_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            context.user_data['act_number'] = update.message.text.strip()
            keyboard = [[
                InlineKeyboardButton("Сегодня", callback_data="DATE:TODAY")
            ]]
            await update.message.reply_text(
                "Какая дата? (введите ДД-ММ-ГГГГ или нажмите 'Сегодня')",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return self.DATE

        async def date_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
            query = update.callback_query
            await query.answer()
            from datetime import datetime as dt
            context.user_data['pour_date'] = dt.now().strftime("%d-%m-%Y")
            return await finalize_and_save(update, context, edit_message=True)

        async def date_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
            text = update.message.text.strip()
            if not re.match(r"^\d{2}-\d{2}-\d{4}$", text):
                await update.message.reply_text("Формат даты: ДД-ММ-ГГГГ")
                return self.DATE
            context.user_data['pour_date'] = text
            return await finalize_and_save(update, context, edit_message=False)

        async def finalize_and_save(update: Update, context: ContextTypes.DEFAULT_TYPE, edit_message: bool):
            data = context.user_data
            # Required fields defaulting
            fields = {
                'object_id': int(data.get('object_id')),
                'pour_date': data.get('pour_date') or '',
                'element': data.get('element') or '',
                'concrete_class': data.get('concrete_class') or '',
                'frost_resistance': data.get('frost_resistance') or '',
                'water_resistance': data.get('water_resistance') or '',
                'supplier': data.get('supplier') or '',
                'concrete_passport': data.get('concrete_passport') or '',
                'volume_concrete': float(data.get('volume_concrete') or 0),
                'cubes_count': int(data.get('cubes_count') or 0),
                'cones_count': int(data.get('cones_count') or 0),
                'slump': data.get('slump') or '',
                'temperature': data.get('temperature') or '',
                'temp_measurements': int(data.get('temp_measurements') or 0),
                'executor': data.get('executor') or '',
                'act_number': data.get('act_number') or '',
                'request_number': '',
                'invoice': ''
            }
            try:
                cur = db_conn.cursor()
                cur.execute(
                    """
                    INSERT INTO constructions (
                        object_id, pour_date, element, concrete_class, frost_resistance,
                        water_resistance, supplier, concrete_passport, volume_concrete, cubes_count,
                        cones_count, slump, temperature, temp_measurements,
                        executor, act_number, request_number, invoice
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        fields['object_id'], fields['pour_date'], fields['element'], fields['concrete_class'],
                        fields['frost_resistance'], fields['water_resistance'], fields['supplier'], fields['concrete_passport'],
                        fields['volume_concrete'], fields['cubes_count'], fields['cones_count'], fields['slump'],
                        fields['temperature'], fields['temp_measurements'], fields['executor'], fields['act_number'],
                        fields['request_number'], fields['invoice']
                    )
                )
                db_conn.commit()
            except Exception as e:
                msg = f"Ошибка сохранения: {str(e)}"
                if edit_message and update.callback_query:
                    await update.callback_query.edit_message_text(msg)
                else:
                    await update.effective_message.reply_text(msg)
                return ConversationHandler.END

            final_msg = "Молодец что не был ленивой жопой и заполнил базу данных"
            if edit_message and update.callback_query:
                await update.callback_query.edit_message_text(final_msg)
            else:
                await update.effective_message.reply_text(final_msg)
            context.user_data.clear()
            return ConversationHandler.END

        async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
            context.user_data.clear()
            await update.effective_message.reply_text("Отменено")
            return ConversationHandler.END

        conv = ConversationHandler(
            entry_points=[CommandHandler("start", start_cmd)],
            states={
                self.ACTION: [CallbackQueryHandler(action_selected, pattern=r"^ACTION:")],
                self.ORG: [CallbackQueryHandler(org_selected, pattern=r"^ORG:")],
                self.OBJ: [CallbackQueryHandler(obj_selected, pattern=r"^OBJ:")],
                self.CLASS: [CallbackQueryHandler(class_selected, pattern=r"^CLASS:")],
                self.FROST: [CallbackQueryHandler(frost_selected, pattern=r"^FROST:")],
                self.WATER: [CallbackQueryHandler(water_selected, pattern=r"^WATER:")],
                self.ELEMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, element_input)],
                self.SUPPLIER: [CallbackQueryHandler(supplier_selected, pattern=r"^SUPPLIER:")],
                self.PASSPORT: [MessageHandler(filters.TEXT & ~filters.COMMAND, passport_input)],
                self.CUBES: [MessageHandler(filters.TEXT & ~filters.COMMAND, cubes_input)],
                self.CONES: [MessageHandler(filters.TEXT & ~filters.COMMAND, cones_input)],
                self.SLUMP: [MessageHandler(filters.TEXT & ~filters.COMMAND, slump_input)],
                self.VOLUME: [MessageHandler(filters.TEXT & ~filters.COMMAND, volume_input)],
                self.TEMP: [MessageHandler(filters.TEXT & ~filters.COMMAND, temp_input)],
                self.TEMP_MEAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, temp_meas_input)],
                self.EXECUTOR: [CallbackQueryHandler(executor_selected, pattern=r"^EXECUTOR:")],
                self.ACT: [MessageHandler(filters.TEXT & ~filters.COMMAND, act_input)],
                self.DATE: [
                    CallbackQueryHandler(date_today, pattern=r"^DATE:TODAY$"),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, date_input)
                ],
            },
            fallbacks=[CommandHandler("cancel", cancel)],
            allow_reentry=True,
        )

        application.add_handler(conv)
        try:
            print("[TG] Preparing event loop...")
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            print("[TG] Starting polling...")
            application.run_polling()
        except Exception as e:
            print(f"[TG] run_polling error: {e}")


class ConcreteApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.db = ConcreteDatabase()
        # Проверяем соединение с базой
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        print("Доступные таблицы:", tables)  # Для отладки
            
        self.current_org_id = None
        self.current_object_id = None
        self.buttons_dict = {}
              
      

        self.title("Учет бетонных работ (prototipe by Kliapich v2.0)")
        self.geometry("1500x600")
        # Стиль ttkbootstrap: тема по умолчанию
        self.style = Style(theme="flatly")
        # Инициализация базы данных
        self.db = ConcreteDatabase()
        self.current_org_id = None
        self.current_object_id = None
        self.buttons_dict = {}
        
        # Настройки панелей
        self.left_panel_width = 180
        self.min_left_panel_width = 150
        self.max_left_panel_width = 350
        
        # Настройки панели объектов
        self.object_panel_height = 200
        self.object_panel_min_height = 100
        self.object_panel_max_height = 400
        
        self.resizing_panel = False
        self.resizing_object_panel = False
        
        # Остальная инициализация...
        self.create_widgets()
        self.load_organizations()
        self.update_buttons_state()

    

    def create_widgets(self):
        ###################### Главные фреймы ############################
        self.left_panel = ttk.Frame(self, width=self.left_panel_width)
        self.left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=0, pady=5)
        self.left_panel.pack_propagate(False)
        
        # Разделители для левой панели
        self.panel_splitter = ttk.Separator(self, orient=tk.VERTICAL)
        self.panel_splitter.pack(side=tk.LEFT, fill=tk.Y, padx=2)
        self.panel_splitter.bind("<Button-1>", self.start_resize)
        self.panel_splitter.bind("<B1-Motion>", self.resize_panel)
        self.panel_splitter.bind("<ButtonRelease-1>", self.stop_resize)
        
        # Правая панель
        right_panel = ttk.Frame(self)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Верхняя часть правой панели (объекты)
        object_frame_container = ttk.Frame(right_panel)
        object_frame_container.pack(fill=tk.X, pady=(0, 5))
        
        # Панель объектов с возможностью изменения высоты
        self.object_frame = ttk.LabelFrame(object_frame_container, text="Объекты", padding=10)
        self.object_frame.pack(fill=tk.X)
        
        # Разделитель для панели объектов
        self.object_panel_splitter = ttk.Separator(object_frame_container, orient=tk.HORIZONTAL)
        self.object_panel_splitter.pack(fill=tk.X, pady=2)
        self.object_panel_splitter.bind("<Button-1>", self.start_resize_object_panel)
        self.object_panel_splitter.bind("<B1-Motion>", self.resize_object_panel)
        self.object_panel_splitter.bind("<ButtonRelease-1>", self.stop_resize_object_panel)
        
        # Таблица объектов
        self.object_tree = ttk.Treeview(self.object_frame, columns=("name", "address"), show="headings", height=5)
        self.object_tree.heading("name", text="Название")
        self.object_tree.heading("address", text="Адрес")
        self.object_tree.bind("<<TreeviewSelect>>", self.on_object_select)
        
        scrollbar_y = ttk.Scrollbar(self.object_frame, orient="vertical", command=self.object_tree.yview)
        scrollbar_x = ttk.Scrollbar(self.object_frame, orient="horizontal", command=self.object_tree.xview)
        self.object_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.object_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        self.object_frame.grid_rowconfigure(0, weight=1)
        self.object_frame.grid_columnconfigure(0, weight=1)
        
        # Нижняя часть правой панели (конструктивы)
        construction_frame = ttk.LabelFrame(right_panel, text="Контроль", padding=10)
        construction_frame.pack(fill=tk.BOTH, expand=True)
        
    
        #################### Кнопки управления ##########################
        # Компактный стиль для кнопок действий
        self.style.configure("Slim.TButton", padding=(6, 2))

        btn_frame = ttk.LabelFrame(self.left_panel, text="Действия", padding=3)
        btn_frame.pack(fill=tk.X, pady=2)

        # кнопки
        buttons = [
            ("+ Организацию", self.add_organization),
            ("- Организацию", self.delete_organization),
            ("Ред. Организацию", self.edit_organization),
            ("+ Объект", self.add_object),
            ("- Объект", self.delete_object),
            ("Ред. Объект", self.edit_object),
            ("+ Контроль", self.add_construction),
            ("- Контроль", self.delete_construction),
            ("Ред. Контроль", self.edit_construction),
            ("Обновить", self.refresh_data),
            ("Создать заявку", self.create_request),
            ("Создать акт", self.create_act),
            ("Шаблон для импорта", self.generate_import_template),
            ("Импорт из Excel", self.import_from_excel),
            ("Экспорт в Excel", self.export_to_excel),
            ("Запустить бота", self.start_telegram_bot)
        ]

        for text, cmd in buttons:
            btn = ttk.Button(btn_frame, text=text, command=cmd, style="Slim.TButton")
            btn.pack(fill=tk.X, pady=1)
            self.buttons_dict[text] = btn

        
        #################### Таблица организаций #########################
        org_frame = ttk.LabelFrame(self.left_panel, text="Организации", padding=5)
        org_frame.pack(fill=tk.BOTH, expand=True)
        
        self.org_tree = ttk.Treeview(org_frame, columns=("name",), show="headings")
        self.org_tree.heading("name", text="Название")
        self.org_tree.column("name", width=self.left_panel_width-30)
        self.org_tree.pack(fill=tk.BOTH, expand=True)
        self.org_tree.bind("<<TreeviewSelect>>", self.on_org_select)
        
        # Выбор темы оформления (перемещено под панель Организации)
        theme_frame = ttk.LabelFrame(self.left_panel, text="Тема", padding=5)
        theme_frame.pack(fill=tk.X, pady=5)
        theme_values = self.style.theme_names()
        self.theme_combo = ttk.Combobox(theme_frame, values=theme_values, state='readonly', width=8)
        self.theme_combo.set(self.style.theme.name)
        self.theme_combo.pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(theme_frame, text="Применить", command=self.apply_selected_theme).pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Фрейм фильтров
        filter_frame = ttk.Frame(construction_frame)
        filter_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(filter_frame, text="Фильтры:").pack(side=tk.LEFT, padx=5)
        
        filters = [
            ("Дата:", "pour_date"),
            ("Класс бетона:", "concrete_class"),
            ("Исполнитель:", "executor"),
            ("Поставщик:", "supplier"),
            ("Счет:", "invoice")
        ]
        
        self.filter_entries = {}
        for text, name in filters:
            frame = ttk.Frame(filter_frame)
            frame.pack(side=tk.LEFT, padx=5)
            ttk.Label(frame, text=text).pack(side=tk.LEFT)
            entry = ttk.Entry(frame, width=15)
            entry.pack(side=tk.LEFT)
            self.filter_entries[name] = entry
        
        ttk.Button(filter_frame, text="Применить", command=self.apply_filters).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text="Сбросить", command=self.reset_filters).pack(side=tk.LEFT, padx=5)
        
        # Таблица конструктивов
        columns = [
            ("Дата", "pour_date", 80),
            ("Конструктив", "element", 120),
            ("Класс", "concrete_class", 50),
            ("Мороз.", "frost_resistance", 60),
            ("Вода.", "water_resistance", 50),
            ("Поставщик", "supplier", 120),
            ("Паспорт", "concrete_passport", 120),
            ("Объем", "volume_concrete", 50),
            ("Кубики", "cubes_count", 60),
            ("Конусы", "cones_count", 60),
            ("Осадка", "slump", 60),
            ("Темп.", "temperature", 50),
            ("Замеры", "temp_measurements", 70),
            ("Исполнитель", "executor", 120),
            ("№ Акта", "act_number", 80),
            ("№ Заявки", "request_number", 80),
            ("Счет", "invoice", 50)
        ]
        
        # Создаем Treeview с колонкой для чекбоксов
        self.construction_tree = ttk.Treeview(
            construction_frame,
            columns=["selected"] + [col[1] for col in columns],
            show="headings",
            height=15,
            selectmode="extended"
        )

        # Колонка с чекбоксами
        self.construction_tree.heading("selected", text="☐", command=self.toggle_all_checkboxes)
        self.construction_tree.column("selected", width=30, anchor="center")
        self.construction_tree.bind("<<TreeviewSelect>>", lambda e: self.update_counters())

        # Остальные колонки
        for text, col, width in columns:
            self.construction_tree.heading(col, text=text)
            self.construction_tree.column(col, width=width, anchor='center')
            
        # Полосы прокрутки
        scrollbar_y = ttk.Scrollbar(construction_frame, orient="vertical", command=self.construction_tree.yview)
        scrollbar_y.pack(side="right", fill="y")
        self.construction_tree.configure(yscrollcommand=scrollbar_y.set)
        self.construction_tree.pack(fill=tk.BOTH, expand=True)

        # Горизонтальная полоса прокрутки для нижней панели
        scrollbar_x = ttk.Scrollbar(construction_frame, orient="horizontal", command=self.construction_tree.xview)
        self.construction_tree.configure(xscrollcommand=scrollbar_x.set)
        scrollbar_x.pack(side="bottom", fill="x")
        

        # Кнопки управления выделением
        btn_frame = ttk.Frame(construction_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        select_all_btn = ttk.Button(btn_frame, text="Выделить все", 
                                command=self.select_all_constructions)
        select_all_btn.pack(side=tk.LEFT, padx=5)

        deselect_all_btn = ttk.Button(btn_frame, text="Снять выделение", 
                                    command=self.deselect_all_constructions)
        deselect_all_btn.pack(side=tk.LEFT, padx=5)

        # Добавление значения в столбец "Счет" для выбранных записей
        self.invoice_value_var = tk.StringVar()
        self.invoice_entry = ttk.Entry(btn_frame, textvariable=self.invoice_value_var, width=18)
        self.invoice_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Добавить в счет", command=self.add_selected_to_invoice).pack(side=tk.LEFT, padx=5)

        # Привязка обработчика кликов
        self.construction_tree.bind("<Button-1>", self.toggle_checkbox)
        # подсчет выделеных строк
        self.status_var = tk.StringVar()
        self.status_var.set("Готово")
        status_bar = ttk.Label(self, textvariable=self.status_var, 
                      style="Status.TLabel", relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.status_frame = ttk.Frame(self)
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)

        # Счетчик выделенных элементов
        self.selected_count_var = tk.StringVar()
        self.selected_count_var.set("Выделено: 0")
        selected_label = ttk.Label(self.status_frame, textvariable=self.selected_count_var)
        selected_label.pack(side=tk.LEFT)

        # Общий счетчик элементов
        self.total_count_var = tk.StringVar()
        self.total_count_var.set("Всего: 0")
        total_label = ttk.Label(self.status_frame, textvariable=self.total_count_var)
        total_label.pack(side=tk.LEFT, padx=10)

        # Статус выполнения операций
        self.status_var = tk.StringVar()
        self.status_var.set("Готово")
        status_label = ttk.Label(self.status_frame, textvariable=self.status_var)
        status_label.pack(side=tk.RIGHT)

        # обновление счетчиков 
    def update_counters(self):
        """Обновляет счетчики выделенных и всех элементов"""
        selected = len(self.construction_tree.selection())
        total = len(self.construction_tree.get_children())
        
        self.selected_count_var.set(f"Выделено: {selected}")
        self.total_count_var.set(f"Всего: {total}")
        
        # Изменяем цвет в зависимости от количества выделенных
        if selected == 0:
            color = "red"
        elif selected == total:
            color = "green"
        else:
            color = "blue"
        
        selected_label = self.status_frame.winfo_children()[0]
        selected_label.config(foreground=color)
     
    ###### Метод для работы с чекбоксами ######
    def update_selection_status(self):
        """Обновляет статусную строку с количеством выделенных элементов"""
        selected_count = len(self.construction_tree.selection())
        total_count = len(self.construction_tree.get_children())
        self.status_var.set(f"Выделено: {selected_count}/{total_count} записей")

    def toggle_all_checkboxes(self):
        current_state = self.construction_tree.heading("selected", "text")
        new_state = "☐" if current_state == "☑" else "☑"
        
        self.construction_tree.heading("selected", text=new_state)
        
        for item in self.construction_tree.get_children():
            self.construction_tree.set(item, "selected", new_state)
        
        self.update_counters()  # Обновляем счетчики

    def toggle_checkbox(self, event):
        region = self.construction_tree.identify("region", event.x, event.y)
        column = self.construction_tree.identify_column(event.x)
        
        if region == "heading" and column == "#1":
            self.toggle_all_checkboxes()
            return
            
        if region == "cell" and column == "#1":
            item = self.construction_tree.identify_row(event.y)
            current_value = self.construction_tree.set(item, "selected")
            new_value = "☑" if current_value != "☑" else "☐"
            self.construction_tree.set(item, "selected", new_value)
        
        self.update_counters()  # Обновляем счетчики
        self.update_header_checkbox_state()

    def select_all_constructions(self):
        items = self.construction_tree.get_children()
        self.construction_tree.selection_set(items)
        for item in items:
            self.construction_tree.set(item, "selected", "☑")
        self.construction_tree.heading("selected", text="☑")
        self.update_counters()  # Обновляем счетчики

    def deselect_all_constructions(self):
        self.construction_tree.selection_remove(self.construction_tree.selection())
        for item in self.construction_tree.get_children():
            self.construction_tree.set(item, "selected", "☐")
        self.construction_tree.heading("selected", text="☐")
        self.update_counters()  # Обновляем счетчики

    def update_header_checkbox_state(self):
        """Обновляет состояние заголовка"""
        selected_count = len(self.construction_tree.selection())
        total_count = len(self.construction_tree.get_children())
        
        if selected_count == 0:
            self.construction_tree.heading("selected", text="☐")
        elif selected_count == total_count:
            self.construction_tree.heading("selected", text="☑")
        else:
            self.construction_tree.heading("selected", text="☒") 
            self.update_selection_status()  

    def add_selected_to_invoice(self):
        """Проставляет указанное значение в поле 'invoice' для выбранных записей."""
        value = self.invoice_value_var.get().strip()
        if not value:
            messagebox.showwarning("Внимание", "Введите значение счета")
            return
        items = self.get_selected_constructions()
        if not items:
            messagebox.showwarning("Внимание", "Выберите хотя бы одну запись")
            return
        try:
            cursor = self.db.conn.cursor()
            placeholders = ','.join(['?'] * len(items))
            params = [value] + list(items)
            cursor.execute(f"UPDATE constructions SET invoice = ? WHERE id IN ({placeholders})", params)
            self.db.conn.commit()
            # Обновляем значения в таблице
            for iid in items:
                if self.construction_tree.exists(iid):
                    self.construction_tree.set(iid, 'invoice', value)
            messagebox.showinfo("Готово", f"Обновлено записей: {cursor.rowcount}")
            # Очистка поля и фокус обратно на ввод
            self.invoice_value_var.set("")
            try:
                self.invoice_entry.focus_set()
            except Exception:
                pass
        except sqlite3.Error as e:
            messagebox.showerror("Ошибка", f"Не удалось обновить счет: {str(e)}")

        
    

    ################## Методы для работы с интерфейсом ##################
    def start_resize(self, event):
        self.resizing_panel = True
        self.start_x = event.x_root
    
    def resize_panel(self, event):
        if self.resizing_panel:
            delta = event.x_root - self.start_x
            new_width = self.left_panel_width + delta
            new_width = max(self.min_left_panel_width, min(self.max_left_panel_width, new_width))
            self.left_panel.config(width=new_width)
            self.org_tree.column("name", width=new_width-30)
            self.left_panel_width = new_width
            self.start_x = event.x_root
 
    def stop_resize(self, event):
        self.resizing_panel = False

    def start_resize_object_panel(self, event):
        self.resizing_object_panel = True
        self.start_y = event.y_root
    
    def resize_object_panel(self, event):
        if self.resizing_object_panel:
            delta = event.y_root - self.start_y
            new_height = self.object_frame.winfo_height() + delta
            new_height = max(self.object_panel_min_height, min(self.object_panel_max_height, new_height))
            self.object_frame.config(height=new_height)
            self.start_y = event.y_root
    
    def stop_resize_object_panel(self, event):
        self.resizing_object_panel = False

    

    def get_selected_constructions(self):
        """Получаем список выбранных конструктивов"""
        selected = []
        for item in self.construction_tree.get_children():
            if self.construction_tree.set(item, "selected") == "☑":
                selected.append(item)
        return selected if selected else self.construction_tree.selection()

    ################## Методы для работы с данными ######################
    def load_organizations(self):
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT id, name FROM organizations")
        self.org_tree.delete(*self.org_tree.get_children())
        for row in cursor.fetchall():
            self.org_tree.insert("", tk.END, values=(row[1],), iid=row[0])

    def on_org_select(self, event=None):
        selected = self.org_tree.selection()
        if selected:
            self.current_org_id = int(selected[0])
            self.load_objects()
        else:
            self.current_org_id = None
        self.current_object_id = None
        self.update_buttons_state()

    def load_objects(self):
        if not self.current_org_id:
            return
            
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT id, name, address FROM objects WHERE org_id = ?", (self.current_org_id,))
        self.object_tree.delete(*self.object_tree.get_children())
        for row in cursor.fetchall():
            self.object_tree.insert("", tk.END, values=row[1:], iid=row[0])

    def on_object_select(self, event=None):
        selected = self.object_tree.selection()
        if selected:
            self.current_object_id = int(selected[0])
            self.load_constructions()
        else:
            self.current_object_id = None
        self.update_buttons_state()

    def load_constructions(self, filters=None):
        if not self.current_object_id:
            return
        
        self.construction_tree.delete(*self.construction_tree.get_children())

        cursor = self.db.conn.cursor()
        query = """
            SELECT id, pour_date, element, concrete_class, frost_resistance, water_resistance,
            supplier, concrete_passport, volume_concrete, cubes_count, cones_count,
            slump, temperature, temp_measurements, executor, act_number, request_number, invoice
            FROM constructions 
            WHERE object_id = ?
        """
        params = [self.current_object_id]

        if filters:
            query += " AND " + " AND ".join([f"{key} LIKE ?" for key in filters.keys()])
            params.extend([f"%{value}%" for value in filters.values()])

        cursor.execute(query, params)
    
        for row in cursor.fetchall():
            values = ["☐"] + [str(row[i]) if row[i] is not None else "" for i in range(1, len(row))]
            self.construction_tree.insert("", "end", values=values, iid=row[0])
            self.update_counters()  # Обновляем счетчики после загрузки
        self.update_header_checkbox_state()
        
    def apply_filters(self):
        filters = {}
        for key, entry in self.filter_entries.items():
            if entry.get():
                filters[key] = entry.get()
        
        self.load_constructions(filters)
    
    def reset_filters(self):
        for entry in self.filter_entries.values():
            entry.delete(0, tk.END)
        self.load_constructions()
        self.update_header_checkbox_state()
    
    def refresh_data(self):
        self.load_organizations()
        if self.current_org_id:
            self.load_objects()
        if self.current_object_id:
            self.load_constructions()
        self.update_buttons_state()

    ################## Методы для работы с организациями ################
    def add_organization(self):
        dialog = tk.Toplevel(self)
        dialog.title("Добавить организацию")
        dialog.geometry("280x170")
        
        fields = [
            ("Название:", "name"),
            ("Контакт:", "contact"),
            ("Телефон:", "phone")
        ]
        
        entries = {}
        for i, (label, name) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            entry = ttk.Entry(dialog, width=30)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky='w')
            entries[name] = entry
        
        def save():
            try:
                self.db.conn.execute(
                    "INSERT INTO organizations (name, contact, phone) VALUES (?, ?, ?)",
                    (entries['name'].get(), entries['contact'].get(), entries['phone'].get())
                )
                self.db.conn.commit()
                self.refresh_data()
                dialog.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("Ошибка", "Организация с таким названием уже существует")
        
        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(fields), columnspan=2, pady=10)
    
    def delete_organization(self):
        selected = self.org_tree.selection()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите организацию для удаления")
            return
        
        if messagebox.askyesno("Подтверждение", "Удалить выбранную организацию и все связанные данные?"):
            self.db.conn.execute("DELETE FROM organizations WHERE id=?", (selected[0],))
            self.db.conn.commit()
            self.current_org_id = None
            self.current_object_id = None
            self.refresh_data()
    
    def edit_organization(self):
        selected = self.org_tree.selection()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите организацию для редактирования")
            return
            
        org_id = int(selected[0])
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT name, contact, phone FROM organizations WHERE id=?", (org_id,))
        org_data = cursor.fetchone()
        
        dialog = tk.Toplevel(self)
        dialog.title("Редактировать организацию")
        dialog.geometry("280x170")
        
        fields = [
            ("Название:", "name", org_data[0]),
            ("Контакт:", "contact", org_data[1]),
            ("Телефон:", "phone", org_data[2])
        ]
        
        entries = {}
        for i, (label, name, value) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            entry = ttk.Entry(dialog, width=30)
            entry.insert(0, value if value else "")
            entry.grid(row=i, column=1, padx=5, pady=5, sticky='w')
            entries[name] = entry
        
        def save():
            try:
                self.db.conn.execute(
                    "UPDATE organizations SET name=?, contact=?, phone=? WHERE id=?",
                    (entries['name'].get(), entries['contact'].get(), entries['phone'].get(), org_id)
                )
                self.db.conn.commit()
                self.refresh_data()
                dialog.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("Ошибка", "Организация с таким названием уже существует")
        
        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(fields), columnspan=2, pady=10)

    ################## Методы для работы с объектами ####################
    def add_object(self):
        if not self.current_org_id:
            messagebox.showwarning("Ошибка", "Сначала выберите организацию")
            return
            
        dialog = tk.Toplevel(self)
        dialog.title("Добавить объект")
        dialog.geometry("220x130")
        
        fields = [
            ("Название:", "name"),
            ("Адрес:", "address")
        ]
        
        entries = {}
        for i, (label, name) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            entry = ttk.Entry(dialog)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky='w')
            entries[name] = entry
        
        def save():
            try:
                self.db.conn.execute(
                    "INSERT INTO objects (org_id, name, address) VALUES (?, ?, ?)",
                    (self.current_org_id, entries['name'].get(), entries['address'].get())
                )
                self.db.conn.commit()
                self.refresh_data()
                dialog.destroy()
            except sqlite3.Error as e:
                messagebox.showerror("Ошибка", str(e))
        
        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(fields), columnspan=2, pady=10)
    
    def delete_object(self):
        selected = self.object_tree.selection()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите объект для удаления")
            return
        
        if messagebox.askyesno("Подтверждение", "Удалить выбранный объект и все связанные данные?"):
            self.db.conn.execute("DELETE FROM objects WHERE id=?", (selected[0],))
            self.db.conn.commit()
            self.current_object_id = None
            self.refresh_data()

    def edit_object(self):
        selected = self.object_tree.selection()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите объект для редактирования")
            return
            
        object_id = int(selected[0])
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT name, address FROM objects WHERE id=?", (object_id,))
        obj_data = cursor.fetchone()
        
        dialog = tk.Toplevel(self)
        dialog.title("Редактировать объект")
        dialog.geometry("220x130")
        
        fields = [
            ("Название:", "name", obj_data[0]),
            ("Адрес:", "address", obj_data[1])
        ]
        
        entries = {}
        for i, (label, name, value) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            entry = ttk.Entry(dialog)
            entry.insert(0, value if value else "")
            entry.grid(row=i, column=1, padx=5, pady=5, sticky='w')
            entries[name] = entry
        
        def save():
            try:
                self.db.conn.execute(
                    "UPDATE objects SET name=?, address=? WHERE id=?",
                    (entries['name'].get(), entries['address'].get(), object_id)
                )
                self.db.conn.commit()
                self.refresh_data()
                dialog.destroy()
            except sqlite3.Error as e:
                messagebox.showerror("Ошибка", str(e))
        
        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(fields), columnspan=2, pady=10)

    ################## Методы для работы с контролем ################
    def add_construction(self):
        if not self.current_object_id:
            messagebox.showwarning("Ошибка", "Сначала выберите объект")
            return
            
        dialog = tk.Toplevel(self)
        dialog.title("Добавить контроль")
        dialog.geometry("280x620")
        dialog.grab_set()

        saved_successfully = False
        
        fields = [
            ("Дата (ДД-ММ-ГГГГ):", "pour_date"),
            ("Конструктив:", "element"),
            ("Класс бетона:", "concrete_class"),
            ("Морозостойкость:", "frost_resistance"),
            ("Водопроницаемость:", "water_resistance"),
            ("Поставщик:", "supplier"),
            ("Паспорт:", "concrete_passport"),
            ("Объем бетона (м³):", "volume_concrete"),
            ("Кубики:", "cubes_count"),
            ("Конусы:", "cones_count"),
            ("Осадка (см):", "slump"),
            ("Температура (°C):", "temperature"),
            ("Замеры темп.:", "temp_measurements"),
            ("Исполнитель:", "executor"),
            ("№ Акта:", "act_number"),
            ("№ Заявки:", "request_number"),
            ("Счет:", "invoice")
        ]
        
        entries = {}
        supplier_values = self._get_distinct_suppliers()
        class_values = ["B7.5", "B10", "B12.5", "B15", "B20", "B22.5", "B25", "B27.5", "B30", "B35", "B40"]
        frost_values = ["F50", "F75", "F100", "F150", "F200", "F300", "F400", "F500"]
        water_values = ["W2", "W4", "W6", "W8", "W10", "W12", "W14"]

        for i, (label, name) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=i, column=0, padx=5, pady=2, sticky='e')
            if name == 'pour_date':
                entry = DateEntry(dialog, firstweekday=0, bootstyle="secondary", dateformat="%d-%m-%Y", width=15)
            elif name == 'supplier':
                entry = ttk.Combobox(dialog, values=supplier_values, state='readonly', width=16)
            elif name == 'concrete_class':
                entry = ttk.Combobox(dialog, values=class_values, state='readonly', width=16)
            elif name == 'frost_resistance':
                entry = ttk.Combobox(dialog, values=frost_values, state='readonly', width=16)
            elif name == 'water_resistance':
                entry = ttk.Combobox(dialog, values=water_values, state='readonly', width=16)
            else:
                entry = ttk.Entry(dialog, width=18)
            entry.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            entries[name] = entry
        
        # Установка значений по умолчанию
        if isinstance(entries['pour_date'], DateEntry):
            entries['pour_date'].set_date(datetime.now())
        else:
            entries['pour_date'].insert(0, datetime.now().strftime("%d-%m-%Y"))
        if isinstance(entries['concrete_class'], ttk.Combobox) and entries['concrete_class']['values']:
            entries['concrete_class'].current(0)
        if isinstance(entries['frost_resistance'], ttk.Combobox) and entries['frost_resistance']['values']:
            entries['frost_resistance'].current(0)
        if isinstance(entries['water_resistance'], ttk.Combobox) and entries['water_resistance']['values']:
            entries['water_resistance'].current(0)
        
        def save():
            try:
                if not entries['pour_date'].get() or not entries['concrete_class'].get():
                    raise ValueError("Заполните обязательные поля (Дата и Класс бетона)")
            
                self.db.conn.execute(
                    """INSERT INTO constructions (
                        object_id, pour_date, element, concrete_class, frost_resistance,
                        water_resistance, supplier, concrete_passport, volume_concrete, cubes_count,
                        cones_count, slump, temperature, temp_measurements,
                        executor, act_number, request_number, invoice
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        self.current_object_id,
                        entries['pour_date'].entry.get() if isinstance(entries['pour_date'], DateEntry) else entries['pour_date'].get(),
                        entries['element'].get(),
                        entries['concrete_class'].get(),
                        entries['frost_resistance'].get(),
                        entries['water_resistance'].get(),
                        entries['supplier'].get(),
                        entries['concrete_passport'].get(),
                        float(entries['volume_concrete'].get() or 0),
                        int(entries['cubes_count'].get() or 0),
                        int(entries['cones_count'].get() or 0),
                        entries['slump'].get(),
                        entries['temperature'].get(),
                        int(entries['temp_measurements'].get() or 0),
                        entries['executor'].get(),
                        entries['act_number'].get(),
                        entries['request_number'].get(),
                        entries['invoice'].get()
                    )
                )
                self.db.conn.commit()
                saved_successfully = True
                dialog.destroy()
                self.refresh_data()
            except ValueError as e:
                messagebox.showerror("Ошибка", str(e))
            except sqlite3.Error as e:
                messagebox.showerror("Ошибка базы данных", str(e))
        
        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(fields), columnspan=2, pady=10)
    
        def on_close():
            if not saved_successfully:
                self.buttons_dict["+ Контроль"]['state'] = 'normal'
            dialog.destroy()
    
        dialog.protocol("WM_DELETE_WINDOW", on_close)

        ############ удаление данных контроля ############
    def delete_construction(self):
        selected = self.get_selected_constructions()  # Используем метод для получения выбранных записей
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите хотя бы один контроль для удаления")
            return
        
        # Подтверждение удаления
        confirm = messagebox.askyesno(
            "Подтверждение удаления",
            f"Вы действительно хотите удалить {len(selected)} выбранных записей?",
            parent=self
        )
        
        if not confirm:
            return
            
        try:
            cursor = self.db.conn.cursor()
            # Формируем параметризованный запрос для удаления нескольких записей
            placeholders = ','.join(['?'] * len(selected))
            query = f"DELETE FROM constructions WHERE id IN ({placeholders})"
            
            cursor.execute(query, selected)
            self.db.conn.commit()
            
            messagebox.showinfo(
                "Успех",
                f"Удалено {cursor.rowcount} записей",
                parent=self
            )
            
            self.refresh_data()
            
        except sqlite3.Error as e:
            messagebox.showerror(
                "Ошибка базы данных",
                f"Не удалось удалить записи:\n{str(e)}",
                parent=self
            )
            ######### конец удаления ###########

    def edit_construction(self):
        selected = self.construction_tree.selection()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите Контроль для редактирования")
            return
            
        constr_id = int(selected[0])
        cursor = self.db.conn.cursor()
        cursor.execute("""
            SELECT pour_date, element, concrete_class, frost_resistance, water_resistance,
                   supplier, concrete_passport, volume_concrete, cubes_count, cones_count,
                   slump, temperature, temp_measurements, executor, act_number, request_number, invoice
            FROM constructions WHERE id=?
        """, (constr_id,))
        constr_data = cursor.fetchone()
        
        dialog = tk.Toplevel(self)
        dialog.title("Редактировать контроль")
        dialog.geometry("280x620")
        dialog.grab_set()
        
        fields = [
            ("Дата (ДД-ММ-ГГГГ):", "pour_date", constr_data[0]),
            ("Конструктив:", "element", constr_data[1]),  
            ("Класс бетона:", "concrete_class", constr_data[2]),
            ("Морозостойкость:", "frost_resistance", constr_data[3]),
            ("Водопроницаемость:", "water_resistance", constr_data[4]),
            ("Поставщик:", "supplier", constr_data[5]),
            ("Паспорт:", "concrete_passport", constr_data[6]),
            ("Объем бетона (м³):", "volume_concrete", constr_data[7]),
            ("Кубики:", "cubes_count", constr_data[8]),
            ("Конусы:", "cones_count", constr_data[9]),
            ("Осадка (см):", "slump", constr_data[10]),
            ("Температура (°C):", "temperature", constr_data[11]),
            ("Замеры темп.:", "temp_measurements", constr_data[12]),
            ("Исполнитель:", "executor", constr_data[13]),
            ("№ Акта:", "act_number", constr_data[14]),
            ("№ Заявки:", "request_number", constr_data[15]),
            ("Счет:", "invoice", constr_data[16])
        ]
        
        entries = {}
        supplier_values = self._get_distinct_suppliers()
        class_values = ["B7.5", "B10", "B12.5", "B15", "B20", "B22.5", "B25", "B27.5", "B30", "B35", "B40"]
        frost_values = ["F50", "F75", "F100", "F150", "F200", "F300", "F400", "F500"]
        water_values = ["W2", "W4", "W6", "W8", "W10", "W12", "W14"]

        for i, (label, name, value) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=i, column=0, padx=5, pady=2, sticky='e')
            if name == 'pour_date':
                entry = DateEntry(dialog, firstweekday=0, bootstyle="secondary", dateformat="%d-%m-%Y", width=15)
                if value:
                    try:
                        entry.set_date(datetime.strptime(str(value), "%d-%m-%Y"))
                    except Exception:
                        try:
                            # Попытка установить напрямую строкой, если формат отличается
                            entry.entry.delete(0, tk.END)
                            entry.entry.insert(0, str(value))
                        except Exception:
                            pass
            elif name == 'supplier':
                entry = ttk.Combobox(dialog, values=supplier_values, state='readonly', width=16)
                if value: entry.set(str(value))
            elif name == 'concrete_class':
                entry = ttk.Combobox(dialog, values=class_values, state='readonly', width=16)
                if value: entry.set(str(value))
            elif name == 'frost_resistance':
                entry = ttk.Combobox(dialog, values=frost_values, state='readonly', width=16)
                if value: entry.set(str(value))
            elif name == 'water_resistance':
                entry = ttk.Combobox(dialog, values=water_values, state='readonly', width=16)
                if value: entry.set(str(value))
            else:
                entry = ttk.Entry(dialog, width=18)
                entry.insert(0, str(value) if value is not None else "")
            entry.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            entries[name] = entry
        
        def save():
            try:
                self.db.conn.execute("""
                    UPDATE constructions SET
                        pour_date=?, element=?, concrete_class=?, frost_resistance=?,
                        water_resistance=?, supplier=?, concrete_passport=?,
                        volume_concrete=?, cubes_count=?, cones_count=?,
                        slump=?, temperature=?, temp_measurements=?,
                        executor=?, act_number=?, request_number=?, invoice=?
                    WHERE id=?
                """, (
                    entries['pour_date'].entry.get() if isinstance(entries['pour_date'], DateEntry) else entries['pour_date'].get(),
                    entries['element'].get(),
                    entries['concrete_class'].get(),
                    entries['frost_resistance'].get(),
                    entries['water_resistance'].get(),
                    entries['supplier'].get(),
                    entries['concrete_passport'].get(),
                    float(entries['volume_concrete'].get() or 0),
                    int(entries['cubes_count'].get() or 0),
                    int(entries['cones_count'].get() or 0),
                    entries['slump'].get(),
                    entries['temperature'].get(),
                    int(entries['temp_measurements'].get() or 0),
                    entries['executor'].get(),
                    entries['act_number'].get(),
                    entries['request_number'].get(),
                    entries['invoice'].get(),
                    constr_id
                ))
                self.db.conn.commit()
                self.refresh_data()
                dialog.destroy()
            except ValueError as e:
                messagebox.showerror("Ошибка", f"Некорректные данные: {str(e)}")
            except sqlite3.Error as e:
                messagebox.showerror("Ошибка базы данных", str(e))
        
        ttk.Button(dialog, text="Сохранить", command=save).grid(row=len(fields), columnspan=2, pady=10)

    ################## Методы для работы с документами ##################
    def create_request(self):
        selected = self.get_selected_constructions()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите хотя бы один контроль")
            return

        for constr_id in selected:
            self.generate_document(constr_id, "request_template.docx", "Заявка")

    def create_act(self):
        selected = self.get_selected_constructions()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите хотя бы один контроль")
            return

        for constr_id in selected:
            self.generate_document(constr_id, "act_template.docx", "Акт")

    def generate_document(self, constr_id, template_name, doc_type):
        """Генерация документа (акта или заявки) на основе шаблона"""
        try:
            # Проверка существования шаблона
            if not os.path.exists(template_name):
                messagebox.showerror("Ошибка", f"Шаблон {template_name} не найден")
                return
            
            cursor = self.db.conn.cursor()
            cursor.execute("""
                SELECT 
                    c.pour_date, c.element, c.concrete_class, c.frost_resistance, c.water_resistance,  
                    c.supplier, c.concrete_passport, c.volume_concrete, c.cubes_count, c.act_number, c.request_number, c.invoice,
                    o.name as object_name, o.address,
                    org.name as org_name, org.contact, org.phone
                FROM constructions c
                JOIN objects o ON c.object_id = o.id
                JOIN organizations org ON o.org_id = org.id
                WHERE c.id = ?
            """, (constr_id,))

            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()
            if not row:
                messagebox.showerror("Ошибка", "Контроль не найден")
                return
            
            construction_data = dict(zip(columns, row))

            # Обработка даты
            #try:
            #     pour_date = datetime.strptime(construction_data['pour_date'], "%d-%m-%Y")
            #   formatted_date = pour_date.strftime("%d.%m.%Y")
            #    file_date = pour_date.strftime(formatted_date)
            #except ValueError:
            #    file_date = "pour_date"
                
                
                ################ Индексы для замены слов ####################
            context = {
                'doc_type': doc_type,
                'current_date': datetime.now().strftime("%d.%m.%Y"),
                'construction': {
                    'object': construction_data.get('object_name', '') or '',
                    'address': construction_data.get('address', '') or '',
                    'date': construction_data.get('pour_date', '') or '',
                    'element': construction_data.get('element', '') or '',  
                    'concrete': construction_data.get('concrete_class', '') or '',
                    'frost': construction_data.get('frost_resistance', '') or '',
                    'water': construction_data.get('water_resistance', '') or '',
                    'supplier': construction_data.get('supplier', '') or '',
                    'passport': construction_data.get('concrete_passport', '') or '',
                    'volume': construction_data.get('volume_concrete', '') or '',
                    'cubes': construction_data.get('cubes_count', '') or '',
                    'cones': construction_data.get('cones_count', '') or '',
                    'slump': construction_data.get('slump', '') or '',
                    'temp': construction_data.get('temperature', '') or '',
                    'temp_measurements': construction_data.get('temp_measurements', '') or '',
                    'act': construction_data.get('act_number', '') or '',
                    'request': construction_data.get('request_number', '') or '',
                    'invoice': construction_data.get('invoice', '') or '' 
                },
                'organization': {
                    'name': construction_data.get('org_name', '') or '',
                    'contact': construction_data.get('contact', '') or '',
                    'phone': construction_data.get('phone', '') or ''
                }
            }
            
                # Формирование имени файла
            object_name = construction_data.get('object_name', 'объект').replace(' ', '_')
            element_name = construction_data.get('element', 'конструктив').replace(' ', '_')
            pour_date = construction_data.get('pour_date', 'дата заливки').replace(' ', '_')
            
            # Удаляем запрещенные символы в имени файла
            safe_object = "".join(c for c in object_name if c.isalnum() or c in (' ', '_')).strip()
            safe_element = "".join(c for c in element_name if c.isalnum() or c in (' ', '_')).strip()
            
            filename = f"{safe_object}_{safe_element}_{pour_date}_{doc_type}.docx"
            
            # Диалог сохранения файла с предложенным именем
            filepath = filedialog.asksaveasfilename(
                defaultextension=".docx",
                filetypes=[("Документ Word", "*.docx")],
                initialfile=filename,  # Предлагаем сформированное имя
                title=f"Сохранить {doc_type.lower()}"
            )
        
            if filepath:
                # Заполнение и сохранение шаблона
                doc = DocxTemplate(template_name)
                doc.render(context)
                doc.save(filepath)
                messagebox.showinfo("Готово", f"{doc_type} сохранен:\n{os.path.basename(filepath)}")
        
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при создании документа:\n{str(e)}")

    ################## Методы для работы с Excel #######################
    def generate_import_template(self):
        """Создание шаблона для импорта"""
        template_type = simpledialog.askstring(
            "Шаблон импорта",
            "Для чего создать шаблон? (org/obj/con):",
            parent=self
        )

        wb = Workbook()
        ws = wb.active

        if template_type == "org":
            ws.append(["Название организации", "Контактное лицо", "Телефон"])
            filename = "template_org.xlsx"
        elif template_type == "obj":
            ws.append(["Название объекта", "Адрес"])
            filename = "template_obj.xlsx"
        elif template_type == "con":
            ws.append([
                "Дата (ДД-ММ-ГГГГ)", "Конструктив", "Класс бетона", "Морозостойкость", 
                "Водопроницаемость", "Поставщик", "Паспорт", "Объем бетона",
                "Кубики", "Конусы", "Осадка", "Температура", "Замеры темп.",
                "Исполнитель", "№ Акта", "№ Заявки", "Счет"
            ])
            filename = "template_constr.xlsx"
        else:
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=filename,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filepath:
            wb.save(filepath)
            messagebox.showinfo("Успех", f"Шаблон сохранен как {filepath}")

    def import_from_excel(self):
        """Импорт данных из Excel файла"""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filepath:
            return

        try:
            wb = load_workbook(filename=filepath)
            ws = wb.active

            import_type = simpledialog.askstring(
                "Тип импорта",
                "Что импортируем? (org - организации, obj - объекты, con - контроль):",
                parent=self
            )

            if import_type == "org":
                self._import_organizations(ws)
            elif import_type == "obj":
                self._import_objects(ws)
            elif import_type == "con":
                self._import_constructions(ws)
            else:
                messagebox.showwarning("Ошибка", "Неверный тип импорта")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при импорте:\n{str(e)}")
        finally:
            self.refresh_data()

    def _import_organizations(self, worksheet):
        """Импорт организаций из Excel"""
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                self.db.conn.execute(
                    """INSERT OR IGNORE INTO organizations 
                    (name, contact, phone) VALUES (?, ?, ?)""",
                    (row[0], row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None)
                )
        self.db.conn.commit()
        messagebox.showinfo("Успех", f"Импортировано {worksheet.max_row-1} организаций")

    def _import_objects(self, worksheet):
        """Импорт объектов из Excel"""
        if not self.current_org_id:
            raise ValueError("Сначала выберите организацию")

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                self.db.conn.execute(
                    """INSERT OR IGNORE INTO objects 
                    (org_id, name, address) VALUES (?, ?, ?)""",
                    (self.current_org_id, row[0], row[1] if len(row) > 1 else None)
                )
        self.db.conn.commit()
        messagebox.showinfo("Успех", f"Импортировано {worksheet.max_row-1} объектов")

    def _import_constructions(self, worksheet):
        """Импорт конструктивов из Excel"""
        if not self.current_object_id:
            raise ValueError("Сначала выберите объект")

        headers = [cell.value for cell in worksheet[1]]
        imported_count = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            data = {
                'object_id': self.current_object_id,
                'pour_date': row[headers.index('Дата (ДД-ММ-ГГГГ)')] if 'Дата (ДД-ММ-ГГГГ)' in headers else None,
                'element': row[headers.index('Конструктив')] if 'Конструктив' in headers else None,
                'concrete_class': row[headers.index('Класс бетона')] if 'Класс бетона' in headers else None,
                'frost_resistance': row[headers.index('Морозостойкость')] if 'Морозостойкость' in headers else None,
                'water_resistance': row[headers.index('Водопроницаемость')] if 'Водопроницаемость' in headers else None,
                'supplier': row[headers.index('Поставщик')] if 'Поставщик' in headers else None,
                'concrete_passport': row[headers.index('Паспорт')] if 'Паспорт' in headers else None,
                'volume_concrete': row[headers.index('Объем бетона')] if 'Объем бетона' in headers else 0,
                'cubes_count': row[headers.index('Кубики')] if 'Кубики' in headers else 0,
                'cones_count': row[headers.index('Конусы')] if 'Конусы' in headers else 0,
                'slump': row[headers.index('Осадка')] if 'Осадка' in headers else None,
                'temperature': row[headers.index('Температура')] if 'Температура' in headers else None,
                'temp_measurements': row[headers.index('Замеры темп.')] if 'Замеры темп.' in headers else 0,
                'executor': row[headers.index('Исполнитель')] if 'Исполнитель' in headers else None,
                'act_number': row[headers.index('№ Акта')] if '№ Акта' in headers else None,
                'request_number': row[headers.index('№ Заявки')] if '№ Заявки' in headers else None,
                'invoice': row[headers.index('Счет')] if 'Счет' in headers else None
            }

            self.db.conn.execute("""
                INSERT INTO constructions (
                    object_id, pour_date, element, concrete_class, frost_resistance,
                    water_resistance, supplier, concrete_passport, volume_concrete, cubes_count,
                    cones_count, slump, temperature, temp_measurements,
                    executor, act_number, request_number, invoice
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, tuple(data.values()))

            imported_count += 1

        self.db.conn.commit()
        messagebox.showinfo("Успех", f"Импортировано {imported_count} контролей")

    def export_to_excel(self):
        if not self.current_object_id:
            messagebox.showwarning("Ошибка", "Сначала выберите объект")
            return
    
        try:
            cursor = self.db.conn.cursor()
            cursor.execute("""
                SELECT 
                    o.name as object_name, 
                    org.name as org_name
                FROM objects o
                JOIN organizations org ON o.org_id = org.id
                WHERE o.id = ?
            """, (self.current_object_id,))
            result = cursor.fetchone()
        
            if not result:
                messagebox.showwarning("Ошибка", "Не удалось получить данные объекта")
                return
            
            object_name = result[0]
            org_name = result[1]
    
            cursor.execute("""
                SELECT pour_date, element, concrete_class, frost_resistance, water_resistance,
                    supplier, concrete_passport, volume_concrete, cubes_count, cones_count,
                    slump, temperature, temp_measurements, executor, act_number, request_number, invoice
                FROM constructions 
                WHERE object_id = ?
                ORDER BY pour_date
            """, (self.current_object_id,))
            constructions = cursor.fetchall()
    
            if not constructions:
                messagebox.showwarning("Ошибка", "Нет данных для экспорта")
                return
    
            wb = Workbook()
            ws = wb.active
            ws.title = "Бетонные работы"
    
            safe_org_name = "".join(x for x in org_name if x.isalnum() or x in (" ", "_")).strip()[:30]
            safe_object_name = "".join(x for x in object_name if x.isalnum() or x in (" ", "_")).strip()[:30]
            default_filename = f"{safe_org_name}_{safe_object_name}_Контроль_бетона.xlsx"

            ws['A1'] = f"Организация: {org_name}"
            ws['A1'].font = Font(bold=True, size=12)
        
            ws['A2'] = f"Объект: {object_name}"
            ws['A2'].font = Font(bold=True, size=12)

            column_widths = {
                'A': 12, 'B': 14, 'C': 18, 'D': 20, 'E': 15, 
                'F': 15, 'G': 15, 'H': 10, 'I': 10, 'J': 10, 
                'K': 15, 'L': 15, 'M': 15, 'N': 10, 'O': 10
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            headers = [
                "Дата", "Конструктив", "Класс бетона", "Морозостойкость", "Водопроницаемость",
                "Поставщик", "Паспорт", "Объем бетона", "Кубики", "Конусы",
                "Осадка", "Температура", "Замеры темп.", "Исполнитель", 
                "№ Акта", "№ Заявки", "Счет"
            ]
            ws.append(headers)
    
            for cell in ws[3]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
    
            for row in constructions:
                ws.append(row)

            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
    
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
            
                for cell in col[2:]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
    
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename
            )
        
            if filename:
                wb.save(filename)
                messagebox.showinfo("Успех", f"Файл успешно сохранен как {filename}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при экспорте:\n{str(e)}")

    ################## Вспомогательные методы ###########################
    def start_telegram_bot(self):
        """Запускает Telegram-бота в фоновом потоке (один раз)."""
        try:
            if not hasattr(self, '_tg_service'):
                self._tg_service = TelegramBotService(TELEGRAM_BOT_TOKEN)
            if self._tg_service.is_running():
                messagebox.showinfo("Бот", "Бот уже запущен")
                return
            self._tg_service.start()
            messagebox.showinfo("Бот", "Бот запущен. В Telegram отправьте /start")
        except Exception as e:
            messagebox.showerror("Ошибка бота", str(e))

    def apply_selected_theme(self):
        """Применяет выбранную тему ttkbootstrap."""
        try:
            selected = getattr(self, 'theme_combo', None)
            if selected is None:
                return
            theme_name = selected.get()
            if theme_name:
                self.style.theme_use(theme_name)
        except Exception as e:
            messagebox.showerror("Ошибка темы", str(e))

    def _get_distinct_suppliers(self):
        """Возвращает список уникальных поставщиков из БД для выпадающего списка."""
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(
                "SELECT DISTINCT supplier FROM constructions "
                "WHERE supplier IS NOT NULL AND supplier != '' ORDER BY supplier"
            )
            return [row[0] for row in cursor.fetchall()]
        except Exception:
            return []

    def update_buttons_state(self):
        """Обновляет состояние кнопок в зависимости от выбора"""
        states = {
            "+ Объект": 'normal' if self.current_org_id else 'disabled',
            "- Объект": 'normal' if self.current_org_id else 'disabled',
            "Ред. Объект": 'normal' if self.current_org_id else 'disabled',
            "+ Контроль": 'normal' if self.current_object_id else 'disabled',
            "- Контроль": 'normal' if self.current_object_id else 'disabled',
            "Ред. Контроль": 'normal' if self.current_object_id else 'disabled'
        }
        
        org_selected = bool(self.org_tree.selection())
        self.buttons_dict["- Организацию"]['state'] = 'normal' if org_selected else 'disabled'
        self.buttons_dict["Ред. Организацию"]['state'] = 'normal' if org_selected else 'disabled'
        
        for btn_text, state in states.items():
            if btn_text in self.buttons_dict:
                self.buttons_dict[btn_text]['state'] = state

if __name__ == "__main__":
    app = ConcreteApp()
    app.mainloop()